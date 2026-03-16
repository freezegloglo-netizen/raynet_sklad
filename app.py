print("APP FILE LOADED v2")

from fastapi import FastAPI, Form, Request, Cookie
from fastapi.responses import HTMLResponse, RedirectResponse, StreamingResponse, JSONResponse
import psycopg2, os, io, json, datetime
from openpyxl import Workbook
from psycopg2.pool import SimpleConnectionPool
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from io import BytesIO
from fastapi.templating import Jinja2Templates
from fastapi.staticfiles import StaticFiles
from fastapi import UploadFile, File
from fastapi.responses import FileResponse
import os

app = FastAPI()

DATABASE_URL = os.getenv("DATABASE_URL")
db_pool = None

templates = Jinja2Templates(directory="templates")
app.mount("/static", StaticFiles(directory="static"), name="static")

def safe_close(conn, cur=None):
    try:
        if cur:
            cur.close()
    except Exception:
        pass

    try:
        if conn:
            db_pool.putconn(conn, close=False)
    except Exception:
        pass

@app.on_event("shutdown")
def shutdown():
    global db_pool
    try:
        if db_pool:
            db_pool.closeall()
            db_pool = None
    except:
        pass


PASSWORD = "morava"

# ================= DB =================


def get_conn():
    if db_pool is None:
        raise Exception("DB NOT CONNECTED")

    return db_pool.getconn()

def init_db():
    conn = None
    cur = None
    try:
        conn = get_conn()
        cur = conn.cursor()

        cur.execute("""
        CREATE TABLE IF NOT EXISTS products (
            id SERIAL PRIMARY KEY,
            code TEXT,
            name TEXT,
            manufacturer TEXT,
            quantity INTEGER DEFAULT 0,
            min_limit INTEGER DEFAULT 5
        );
        """)

        cur.execute("""
        CREATE TABLE IF NOT EXISTS movements (
            id SERIAL PRIMARY KEY,
            code TEXT,
            change INTEGER,
            user_name TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        );
        """)

        cur.execute("""
        CREATE TABLE IF NOT EXISTS car_stock (
            id SERIAL PRIMARY KEY,
            user_name TEXT,
            code TEXT,
            quantity INTEGER DEFAULT 0,
            UNIQUE(user_name, code)
        );
        """)

        cur.execute("""
        CREATE TABLE IF NOT EXISTS app_users (
            id SERIAL PRIMARY KEY,
            username TEXT UNIQUE NOT NULL,
            display_name TEXT NOT NULL,
            role TEXT DEFAULT 'driver',
            active BOOLEAN DEFAULT TRUE,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        );
        """)

        cur.execute("""
        CREATE TABLE IF NOT EXISTS app_settings (
            id SERIAL PRIMARY KEY,
            login_background TEXT DEFAULT '/static/login_bg.jpg',
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        );
        """)

        cur.execute("""
        INSERT INTO app_users (username, display_name, role)
        VALUES
            ('Lukas', 'Lukáš', 'driver'),
            ('Jirka', 'Jirka', 'driver'),
            ('Milan', 'Milan', 'driver')
        ON CONFLICT (username) DO NOTHING;
        """)

        cur.execute("""
        INSERT INTO app_settings (login_background)
        SELECT '/static/login_bg.jpg'
        WHERE NOT EXISTS (SELECT 1 FROM app_settings);
        """)

        conn.commit()

    except Exception as e:
        print("DB INIT FAILED:", e)

    finally:
        safe_close(conn, cur)

@app.on_event("startup")
def startup():
    global db_pool
    print("INIT DB POOL...")

    try:
        db_pool = SimpleConnectionPool(
            minconn=1,
            maxconn=10,
            dsn=DATABASE_URL + "?sslmode=require"
        )

        print("DB POOL READY")

        init_db()  # ← TADY se vytvoří tabulky

    except Exception as e:
        print("DB INIT FAILED:", e)
        db_pool = None# 

def get_app_users():
    conn = None
    cur = None
    try:
        conn = get_conn()
        cur = conn.cursor()
        cur.execute("""
            SELECT username, display_name, role
            FROM app_users
            WHERE active = TRUE
            ORDER BY id
        """)
        return cur.fetchall()
    finally:
        safe_close(conn, cur)


def get_login_background():
    conn = None
    cur = None
    try:
        conn = get_conn()
        cur = conn.cursor()
        cur.execute("""
            SELECT login_background
            FROM app_settings
            ORDER BY id
            LIMIT 1
        """)
        row = cur.fetchone()
        return row[0] if row else "/static/login_bg.jpg"
    finally:
        safe_close(conn, cur)

# ================= LOGIN =================
@app.get("/login", response_class=HTMLResponse)
def login_page():
    bg = get_login_background()
    return f"""
    <!DOCTYPE html>
    <html lang="cs">
    <head>
        <meta charset="utf-8">
        <meta name="viewport" content="width=device-width, initial-scale=1.0">
        <title>Přihlášení</title>
        <style>
            * {{ box-sizing:border-box; }}
            body {{
                margin:0;
                min-height:100vh;
                font-family:Arial,sans-serif;
                color:#fff;
                background:
                    linear-gradient(rgba(0,0,0,.72), rgba(0,0,0,.82)),
                    url('{bg}') center/cover no-repeat;
                display:flex;
                align-items:center;
                justify-content:center;
            }}
            .card {{
                width:min(420px, 92vw);
                background:rgba(15,15,18,.78);
                border:1px solid rgba(255,255,255,.08);
                border-radius:24px;
                padding:34px;
                backdrop-filter: blur(14px);
                box-shadow:0 20px 60px rgba(0,0,0,.35);
            }}
            h1 {{
                margin:0 0 22px;
                font-size:34px;
            }}
            .sub {{
                color:#b8b8b8;
                margin-bottom:22px;
            }}
            input {{
                width:100%;
                padding:15px 16px;
                border-radius:14px;
                border:1px solid #2a2a2a;
                background:#111;
                color:#fff;
                font-size:16px;
                margin-bottom:14px;
            }}
            button {{
                width:100%;
                padding:15px 16px;
                border:none;
                border-radius:14px;
                background:#18c37e;
                color:#fff;
                font-size:16px;
                font-weight:bold;
                cursor:pointer;
            }}
            button:hover {{
                filter:brightness(1.05);
            }}
        </style>
    </head>
    <body>
        <div class="card">
            <h1>Přihlášení</h1>
            <div class="sub">Vstup do skladového systému</div>
            <form method="post" action="/login">
                <input type="password" name="password" placeholder="Zadej heslo" required>
                <button type="submit">Pokračovat</button>
            </form>
        </div>
    </body>
    </html>
    """

@app.get("/logout")
def logout():
    r = RedirectResponse("/login", status_code=303)
    r.delete_cookie("user")
    r.delete_cookie("mode")
    r.delete_cookie("auth")
    return r

@app.post("/login")
def login(password: str = Form(...)):
    if password == PASSWORD:
        r = RedirectResponse("/select_user", status_code=303)
        r.set_cookie("auth", "ok")
        return r
    return RedirectResponse("/login", status_code=303)

# ================= USER SELECT =================
@app.get("/select_user", response_class=HTMLResponse)
def select_user(auth: str = Cookie(default=None), edit: int = 0):
    if auth != "ok":
        return RedirectResponse("/login", status_code=303)

    users = get_app_users()
    bg = get_login_background()
    is_edit = edit == 1

    html = f"""
    <!DOCTYPE html>
    <html lang="cs">
    <head>
        <meta charset="utf-8">
        <meta name="viewport" content="width=device-width, initial-scale=1.0">
        <title>Výběr uživatele</title>
        <style>
            * {{ box-sizing:border-box; }}
            body {{
                margin:0;
                min-height:100vh;
                font-family:Arial,sans-serif;
                color:#fff;
                background:
                    linear-gradient(rgba(0,0,0,.62), rgba(0,0,0,.86)),
                    url('{bg}') center/cover no-repeat;
            }}

            .topbar {{
                display:flex;
                justify-content:space-between;
                align-items:center;
                padding:24px 34px;
            }}

            .brand {{
                font-size:22px;
                font-weight:700;
                letter-spacing:.4px;
            }}

            .top-actions {{
                display:flex;
                gap:12px;
                align-items:center;
            }}

            .top-btn {{
                text-decoration:none;
                color:#fff;
                background:rgba(255,255,255,.08);
                border:1px solid rgba(255,255,255,.12);
                border-radius:12px;
                padding:10px 16px;
                font-size:14px;
            }}

            .top-btn.primary {{
                background:#18c37e;
                color:#fff;
                border:none;
            }}

            .wrap {{
                min-height:calc(100vh - 90px);
                display:flex;
                justify-content:center;
                align-items:center;
                padding:30px;
            }}

            .panel {{
                width:min(1100px, 100%);
                background:rgba(10,10,12,.50);
                border:1px solid rgba(255,255,255,.08);
                border-radius:30px;
                padding:42px;
                backdrop-filter: blur(8px);
            }}

            h1 {{
                margin:0 0 10px;
                font-size:44px;
                text-align:center;
            }}

            .sub {{
                text-align:center;
                color:#c7c7c7;
                margin-bottom:34px;
                font-size:18px;
            }}

            .grid {{
                display:grid;
                grid-template-columns:repeat(auto-fit, minmax(190px, 1fr));
                gap:22px;
            }}

            .profile-card {{
                background:rgba(255,255,255,.08);
                border:1px solid rgba(255,255,255,.10);
                border-radius:24px;
                min-height:220px;
                padding:22px;
                display:flex;
                flex-direction:column;
                justify-content:space-between;
                transition:.2s ease;
            }}

            .profile-card:hover {{
                transform:translateY(-4px) scale(1.01);
                background:rgba(255,255,255,.12);
            }}

            .avatar {{
                width:78px;
                height:78px;
                border-radius:50%;
                background:linear-gradient(135deg,#18c37e,#1095c1);
                display:flex;
                align-items:center;
                justify-content:center;
                font-size:28px;
                font-weight:bold;
                margin-bottom:18px;
            }}

            .name {{
                font-size:24px;
                font-weight:700;
                margin-bottom:12px;
            }}

            .enter-form button,
            .action-btn {{
                width:100%;
                padding:12px 14px;
                border:none;
                border-radius:14px;
                cursor:pointer;
                font-size:15px;
            }}

            .enter-form button {{
                background:#fff;
                color:#111;
                font-weight:bold;
            }}

            .action-btn {{
                background:#18c37e;
                color:#fff;
                text-decoration:none;
                display:inline-flex;
                align-items:center;
                justify-content:center;
            }}

            .edit-box {{
                margin-top:14px;
                display:flex;
                flex-direction:column;
                gap:10px;
            }}

            .edit-box input {{
                width:100%;
                padding:11px 12px;
                border-radius:12px;
                border:1px solid #2d2d2d;
                background:#111;
                color:#fff;
            }}

            .danger {{
                background:#a92a2a !important;
                color:#fff;
            }}

            .footer-actions {{
                margin-top:28px;
                display:flex;
                gap:14px;
                flex-wrap:wrap;
                justify-content:center;
            }}

            .footer-actions form {{
                display:inline-block;
            }}

            .footer-actions input {{
                padding:12px 14px;
                border-radius:12px;
                border:1px solid #2d2d2d;
                background:#111;
                color:#fff;
                min-width:220px;
            }}

            .warehouse-btn {{
                margin-top:34px;
                text-align:center;
            }}

            .warehouse-btn button {{
                padding:16px 28px;
                font-size:18px;
                border:none;
                border-radius:16px;
                background:#18c37e;
                color:#fff;
                cursor:pointer;
                font-weight:bold;
            }}

            @media (max-width: 700px) {{
                h1 {{ font-size:32px; }}
                .panel {{ padding:24px; }}
            }}
        </style>
    </head>
    <body>
        <div class="topbar">
            <div class="brand">MADISSON</div>
            <div class="top-actions">
                {"<a class='top-btn' href='/select_user'>Hotovo</a>" if is_edit else "<a class='top-btn primary' href='/select_user?edit=1'>Upravit</a>"}
                <a class="top-btn" href="/logout">Odhlásit</a>
            </div>
        </div>

        <div class="wrap">
            <div class="panel">
                <h1>Kdo dnes používá sklad?</h1>
                <div class="sub">Vyber uživatele nebo přepni do režimu úprav</div>

                <div class="grid">
    """

    for username, display_name, role in users:
        initial = (display_name[:1] or "?").upper()

        html += f"""
        <div class="profile-card">
            <div>
                <div class="avatar">{initial}</div>
                <div class="name">{display_name}</div>
            </div>
        """

        if not is_edit:
            html += f"""
            <form method="post" action="/set_user" class="enter-form">
                <input type="hidden" name="user" value="{username}">
                <button type="submit">Vstoupit</button>
            </form>
            """
        else:
            html += f"""
            <div class="edit-box">
                <form method="post" action="/rename_user">
                    <input type="hidden" name="username" value="{username}">
                    <input type="text" name="display_name" value="{display_name}" required>
                    <button type="submit" class="action-btn">Přejmenovat</button>
                </form>

                <form method="post" action="/delete_user">
                    <input type="hidden" name="username" value="{username}">
                    <button type="submit" class="action-btn danger">Odebrat</button>
                </form>
            </div>
            """

        html += "</div>"

    html += """
                </div>
    """

    if is_edit:
        html += """
                <div class="footer-actions">
                    <form method="post" action="/create_user">
                        <input type="text" name="username" placeholder="Interní jméno, např. Tomas" required>
                        <input type="text" name="display_name" placeholder="Zobrazené jméno, např. Tomáš" required>
                        <button class="action-btn" type="submit">Přidat řidiče</button>
                    </form>

                    <form method="post" action="/upload_login_background" enctype="multipart/form-data">
                        <input type="file" name="photo" accept="image/*" required>
                        <button class="action-btn" type="submit">Změnit hlavní fotku</button>
                    </form>
                </div>
        """

    html += """
                <div class="warehouse-btn">
                    <form method="post" action="/set_sklad">
                        <button type="submit">📦 Přehled / Úprava skladu</button>
                    </form>
                </div>
            </div>
        </div>
    </body>
    </html>
    """

    return HTMLResponse(html)

import urllib.parse

@app.post("/set_user")
def set_user(user: str = Form(...)):
    r = RedirectResponse("/", status_code=303)

    import urllib.parse
    safe_user = urllib.parse.quote(user)

    r.set_cookie("user", safe_user)
    r.set_cookie("mode", "driver")

    return r

@app.post("/set_sklad")
def set_sklad():
    r = RedirectResponse("/", status_code=303)
    r.set_cookie("user", "Sklad")
    r.set_cookie("mode", "sklad")
    return r

@app.post("/create_user")
def create_user(
    auth: str = Cookie(default=None),
    username: str = Form(...),
    display_name: str = Form(...)
):
    if auth != "ok":
        return RedirectResponse("/login", status_code=303)

    conn = None
    cur = None
    try:
        conn = get_conn()
        cur = conn.cursor()
        cur.execute("""
            INSERT INTO app_users (username, display_name, role)
            VALUES (%s, %s, 'driver')
            ON CONFLICT (username) DO NOTHING
        """, (username.strip(), display_name.strip()))
        conn.commit()
    finally:
        safe_close(conn, cur)

    return RedirectResponse("/select_user?edit=1", status_code=303)


@app.post("/rename_user")
def rename_user(
    auth: str = Cookie(default=None),
    username: str = Form(...),
    display_name: str = Form(...)
):
    if auth != "ok":
        return RedirectResponse("/login", status_code=303)

    conn = None
    cur = None
    try:
        conn = get_conn()
        cur = conn.cursor()
        cur.execute("""
            UPDATE app_users
            SET display_name=%s
            WHERE username=%s
        """, (display_name.strip(), username))
        conn.commit()
    finally:
        safe_close(conn, cur)

    return RedirectResponse("/select_user?edit=1", status_code=303)


@app.post("/delete_user")
def delete_user(
    auth: str = Cookie(default=None),
    username: str = Form(...)
):
    if auth != "ok":
        return RedirectResponse("/login", status_code=303)

    conn = None
    cur = None
    try:
        conn = get_conn()
        cur = conn.cursor()
        cur.execute("""
            UPDATE app_users
            SET active=FALSE
            WHERE username=%s
        """, (username,))
        conn.commit()
    finally:
        safe_close(conn, cur)

    return RedirectResponse("/select_user?edit=1", status_code=303)

# ================= DASHBOARD =================
@app.get("/", response_class=HTMLResponse)
def home(request: Request, auth: str = Cookie(default=None)):

    if auth != "ok":
        return RedirectResponse("/login", status_code=303)

    conn = None
    cur = None

    try:
        conn = get_conn()
        cur = conn.cursor()

        cur.execute("""
        SELECT manufacturer, SUM(quantity)
        FROM products
        GROUP BY manufacturer
        ORDER BY manufacturer
        """)
        data = cur.fetchall()

        labels = [d[0] or "Neznámý" for d in data]
        values = [int(d[1]) for d in data]

        cur.execute("SELECT COUNT(*) FROM products")
        total_products = cur.fetchone()[0]

        cur.execute("SELECT COUNT(*) FROM products WHERE quantity <= min_limit")
        low_products = cur.fetchone()[0]

        cur.execute("SELECT COUNT(*) FROM movements WHERE DATE(created_at)=CURRENT_DATE")
        today_moves = cur.fetchone()[0]

        cur.execute("SELECT COUNT(DISTINCT manufacturer) FROM products")
        manufacturers = cur.fetchone()[0]

    except Exception as e:
        print("HOME ERROR:", e)

    finally:
        safe_close(conn, cur)

    user = request.cookies.get("user", "Neznámý")
    mode = request.cookies.get("mode", "driver")
    mode_label = "SKLAD" if mode == "sklad" else "ŘIDIČ"

    return templates.TemplateResponse(
        "dashboard.html",
        {
            "request": request,
            "title": "Dashboard",
            "user": user,
            "mode": mode_label,
            "total_products": total_products,
            "low_products": low_products,
            "today_moves": today_moves,
            "manufacturers": manufacturers,
            "labels": labels,
            "values": values
        }
    )

# ================= API =================
@app.get("/api/manufacturers")
def api_man():
    conn=get_conn()
    cur=conn.cursor()
    cur.execute("SELECT DISTINCT manufacturer FROM products ORDER BY manufacturer")
    data=[m[0] or "Neznámý" for m in cur.fetchall()]
    safe_close(conn, cur)
    return data

@app.get("/car", response_class=HTMLResponse)
def car(request: Request, auth: str = Cookie(default=None), user: str = Cookie(default=None)):
    if auth != "ok":
        return RedirectResponse("/login", status_code=303)

    import urllib.parse
    if user:
        user = urllib.parse.unquote(user)

    if not user:
        return RedirectResponse("/select_user", status_code=303)

    conn = get_conn()
    cur = conn.cursor()

    cur.execute("""
        SELECT c.code, p.name, c.quantity
        FROM car_stock c
        JOIN products p ON p.code = c.code
        WHERE c.user_name=%s
        ORDER BY c.code
    """, (user,))
    rows = cur.fetchall()

    html = """
    <html>
    <head>
    <meta charset="utf-8">
    <style>
    body{background:#0f1115;color:#eee;font-family:Arial;margin:0;padding:20px}
    table{width:100%;border-collapse:collapse}
    th,td{padding:10px;border-bottom:1px solid #222}
    button{background:#2b3445;color:#fff;border:none;padding:6px 10px;border-radius:6px;cursor:pointer}
    button:hover{background:#3b4760}
    </style>
    </head>
    <body>
    """

    # ===== HLAVIČKA =====
    mode = request.cookies.get("mode", "driver")
    mode_label = "SKLAD" if mode == "sklad" else "ŘIDIČ"

    html += f"""
    <div style="margin-bottom:15px">
        👤 <b>{user}</b> | Režim: <b>{mode_label}</b>
    </div>

    <a href="/"><button>Zpět</button></a>

    <h2>Auto — {user}</h2>

    <table>
    <tr><th>Kód</th><th>Název</th><th>Množství</th><th>Akce</th></tr>
    """

    if not rows:
        html += "<tr><td colspan=3 style='color:#777'>Prázdné auto</td></tr>"
    else:
        for code, name, qty in rows:
            html += f"""
            <tr>
                <td>{code}</td>
                <td>{name}</td>
                <td>{qty}</td>
                <td>
                    <form method="post" action="/use_from_car" style="display:inline">
                        <input type="hidden" name="code" value="{code}">
                        <button>Použito</button>
                    </form>
                </td>
            </tr>
            """

    html += """
    </table>
    </body>
    </html>
    """

    safe_close(conn, cur)
    return HTMLResponse(html)

@app.get("/cars", response_class=HTMLResponse)
def cars(request: Request, auth: str = Cookie(default=None)):
    if auth != "ok":
        return RedirectResponse("/login", status_code=303)

    import urllib.parse

    conn = None
    cur = None

    try:
        conn = get_conn()
        cur = conn.cursor()

        raw_user = request.cookies.get("user", "Neznámý")
        user = urllib.parse.unquote(raw_user) if raw_user else "Neznámý"
        mode = request.cookies.get("mode", "driver")

        cars = []

        if mode == "sklad":
            users = get_app_users()

            for key, label, role in users:
                cur.execute("""
                    SELECT c.code, p.name, c.quantity
                    FROM car_stock c
                    JOIN products p ON p.code = c.code
                    WHERE c.user_name=%s AND c.quantity > 0
                    ORDER BY c.code
                """, (key,))
                rows = cur.fetchall()

                items = []
                for r in rows:
                    items.append({
                        "code": r[0],
                        "name": r[1],
                        "quantity": r[2]
                    })

                cars.append({
                    "user_key": key,
                    "user": label,
                    "products": items
                })

            page_title = "Auta"

        else:
            cur.execute("""
                SELECT display_name
                FROM app_users
                WHERE username=%s
                LIMIT 1
            """, (user,))
            row = cur.fetchone()
            display_name = row[0] if row else user

            cur.execute("""
                SELECT c.code, p.name, c.quantity
                FROM car_stock c
                JOIN products p ON p.code = c.code
                WHERE c.user_name=%s AND c.quantity > 0
                ORDER BY c.code
            """, (user,))
            rows = cur.fetchall()

            items = []
            for r in rows:
                items.append({
                    "code": r[0],
                    "name": r[1],
                    "quantity": r[2]
                })

            cars.append({
                "user_key": user,
                "user": display_name,
                "products": items
            })

            page_title = "Auto"

    finally:
        safe_close(conn, cur)

    return templates.TemplateResponse(
        "cars_new.html",
        {
            "request": request,
            "title": page_title,
            "user": user,
            "mode": mode,
            "cars": cars
        }
    )

@app.get("/export/products")
def export_products(auth: str = Cookie(default=None)):

    if auth != "ok":
        return RedirectResponse("/login", status_code=303)

    conn = get_conn()
    cur = conn.cursor()

    cur.execute("""
        SELECT code, name, manufacturer, quantity, min_limit
        FROM products
        ORDER BY manufacturer, name
    """)
    rows = cur.fetchall()
    safe_close(conn, cur)

    wb = Workbook()
    wb.remove(wb.active)

    sheets = {}

    for code, name, manufacturer, qty, minl in rows:

        man = manufacturer or "Neznámý"

        if man not in sheets:
            ws = wb.create_sheet(title=man[:31])
            ws.append(["Kód", "Název", "Množství", "Min. limit"])
            sheets[man] = ws

        sheets[man].append([code, name, qty, minl])

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)

    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=produkty.xlsx"}
    )

@app.get("/export/low")
def export_low(auth: str = Cookie(default=None)):

    if auth != "ok":
        return RedirectResponse("/login", status_code=303)

    conn = get_conn()
    cur = conn.cursor()

    cur.execute("""
        SELECT code, name, manufacturer, quantity, min_limit
        FROM products
        WHERE quantity <= min_limit
        ORDER BY manufacturer, name
    """)
    rows = cur.fetchall()
    safe_close(conn, cur)

    wb = Workbook()
    wb.remove(wb.active)

    sheets = {}

    for code, name, manufacturer, qty, minl in rows:

        man = manufacturer or "Neznámý"

        if man not in sheets:
            ws = wb.create_sheet(title=man[:31])
            ws.append(["Kód", "Název", "Množství", "Min. limit"])
            sheets[man] = ws

        sheets[man].append([code, name, qty, minl])

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)

    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=nizky_stav.xlsx"}
    )

@app.get("/api/history/{manufacturer}")
def api_hist(manufacturer:str):
    conn=None
    cur=None
    try:
        conn=get_conn()
        cur=conn.cursor()
        cur.execute("""
        SELECT code,change,created_at FROM movements
        WHERE code IN (SELECT code FROM products WHERE manufacturer=%s)
        ORDER BY created_at
        """,(manufacturer,))
        rows=cur.fetchall()

    finally:
        if cur: cur.close()
        if conn: db_pool.putconn(conn)

    timeline={}
    for code,ch,ts in rows:
        timeline.setdefault(code,{"t":[],"v":[],"s":0})
        timeline[code]["s"]+=ch
        timeline[code]["t"].append(str(ts))
        timeline[code]["v"].append(timeline[code]["s"])
    return timeline

@app.post("/add")
def add(
    code: str = Form(...),
    name: str = Form(...),
    manufacturer: str = Form(...),
    quantity: int = Form(0),
    min_limit: int = Form(5)
):
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("""
        INSERT INTO products(code,name,manufacturer,quantity,min_limit)
        VALUES(%s,%s,%s,%s,%s)
    """, (code,name,manufacturer,quantity,min_limit))
    conn.commit()
    safe_close(conn, cur)
    return RedirectResponse("/all_new", status_code=303)

@app.post("/change")
def change(code: str = Form(...), type: str = Form(...), user: str = Cookie(default=None)):
    import urllib.parse

    if user:
        user = urllib.parse.unquote(user)

    conn = None
    cur = None

    try:
        conn = get_conn()
        cur = conn.cursor()

        cur.execute("SELECT quantity FROM products WHERE code=%s", (code,))
        row = cur.fetchone()
        if not row:
            return RedirectResponse("/all_new", status_code=303)

        qty = row[0]

        if type == "add":
            qty += 1
            change_val = 1
        else:
            qty = max(0, qty - 1)
            change_val = -1

        cur.execute("UPDATE products SET quantity=%s WHERE code=%s", (qty, code))

        cur.execute(
            "INSERT INTO movements(code, change, user_name) VALUES(%s,%s,%s)",
            (code, change_val, user or "Unknown")
        )

        conn.commit()

    except Exception as e:
        print("CHANGE ERROR:", e)
        if conn:
            conn.rollback()
        return HTMLResponse(f"<h1>DB ERROR</h1><pre>{e}</pre>", status_code=500)

    finally:
        if cur:
            cur.close()
        if conn:
            db_pool.putconn(conn)

    return RedirectResponse("/all_new", status_code=303)

@app.post("/delete_by_code")
def delete_by_code(code: str = Form(...)):
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("DELETE FROM products WHERE code=%s", (code,))
    conn.commit()
    safe_close(conn, cur)
    return RedirectResponse("/all_new", status_code=303)

@app.post("/choose_car")
def choose_car(code: str = Form(...), auth: str = Cookie(default=None)):
    if auth != "ok":
        return RedirectResponse("/login", status_code=303)

    users = get_app_users()

    html = """
    <html>
    <head>
    <meta charset="utf-8">
    <style>
    body{
        background:#111;
        color:#eee;
        font-family:Arial;
        text-align:center;
        padding:40px;
    }
    button{
        padding:10px 20px;
        margin:10px;
        border:none;
        border-radius:8px;
        background:#2b5;
        color:#fff;
        font-size:16px;
        cursor:pointer;
    }
    button:hover{
        opacity:0.92;
    }
    </style>
    </head>
    <body>
    <h2>Vyber auto</h2>
    """

    for key, label, role in users:
        html += f"""
        <form method="post" action="/to_car">
            <input type="hidden" name="code" value="{code}">
            <input type="hidden" name="user" value="{key}">
            <button>{label}</button>
        </form>
        """

    html += """
    </body>
    </html>
    """

    return HTMLResponse(html)

@app.post("/use_from_car")
def use_from_car(
    request: Request,
    code: str = Form(...),
    target_user: str = Form(None),
    user: str = Cookie(default=None),
    auth: str = Cookie(default=None)
):
    import urllib.parse

    if auth != "ok":
        return RedirectResponse("/login", status_code=303)

    if user:
        user = urllib.parse.unquote(user)

    mode = request.cookies.get("mode", "driver")
    final_user = target_user if mode == "sklad" and target_user else user

    if not final_user:
        return RedirectResponse("/cars", status_code=303)

    conn = None
    cur = None

    try:
        conn = get_conn()
        cur = conn.cursor()

        cur.execute("""
            UPDATE car_stock
            SET quantity = quantity - 1
            WHERE user_name=%s AND code=%s AND quantity > 0
            RETURNING quantity
        """, (final_user, code))

        if cur.fetchone() is None:
            conn.commit()
            return RedirectResponse("/cars", status_code=303)

        cur.execute("""
            INSERT INTO movements(code, change, user_name)
            VALUES(%s, %s, %s)
        """, (code, -1, final_user))

        conn.commit()

    finally:
        safe_close(conn, cur)

    return RedirectResponse("/cars", status_code=303)

@app.post("/return_from_car")
def return_from_car(
    code: str = Form(...),
    target_user: str = Form(None),
    request: Request = None,
    auth: str = Cookie(default=None)
):
    if auth != "ok":
        return RedirectResponse("/login", status_code=303)

    import urllib.parse

    mode = request.cookies.get("mode", "driver")
    cookie_user = request.cookies.get("user", None)
    if cookie_user:
        cookie_user = urllib.parse.unquote(cookie_user)

    final_user = target_user if mode == "sklad" and target_user else cookie_user

    if not final_user:
        return RedirectResponse("/cars", status_code=303)

    conn = None
    cur = None

    try:
        conn = get_conn()
        cur = conn.cursor()

        cur.execute("""
            UPDATE car_stock
            SET quantity = quantity - 1
            WHERE user_name=%s AND code=%s AND quantity > 0
            RETURNING quantity
        """, (final_user, code))

        if cur.fetchone() is None:
            conn.commit()
            return RedirectResponse("/cars", status_code=303)

        cur.execute("""
            UPDATE products
            SET quantity = quantity + 1
            WHERE code=%s
        """, (code,))

        cur.execute("""
            INSERT INTO movements(code, change, user_name)
            VALUES(%s, %s, %s)
        """, (code, 1, final_user))

        conn.commit()

    finally:
        safe_close(conn, cur)

    return RedirectResponse("/cars", status_code=303)

@app.post("/to_car")
def to_car(code: str = Form(...),
           qty: int = Form(1),
           user: str = Form(None),
           user_cookie: str = Cookie(default=None, alias="user"),
           auth: str = Cookie(default=None)):

    if auth != "ok":
        return RedirectResponse("/login", status_code=303)

    import urllib.parse

    final_user = None

    if user:
        final_user = urllib.parse.unquote(user)
    elif user_cookie:
        final_user = urllib.parse.unquote(user_cookie)

    if not final_user:
        return RedirectResponse("/all_new", status_code=303)
    conn = None
    cur = None

    try:
        conn = get_conn()
        cur = conn.cursor()

        cur.execute("""
            UPDATE products
            SET quantity = quantity - %s
            WHERE code=%s AND quantity >= %s
            RETURNING quantity
        """, (qty, code, qty))

        if cur.fetchone() is None:
            conn.commit()
            return RedirectResponse("/all_new", status_code=303)
                              
        # 🔹 ZÁPIS DO HISTORIE – přesun do auta
        cur.execute("""
            INSERT INTO movements(code, change, user_name)
            VALUES(%s, %s, %s)
        """, (code, -qty, final_user))               
                          
        cur.execute("""
            INSERT INTO car_stock(user_name, code, quantity)
            VALUES(%s,%s,%s)
            ON CONFLICT (user_name, code)
            DO UPDATE SET quantity = car_stock.quantity + %s
        """, (final_user, code, qty, qty))

        conn.commit()

    except Exception as e:
        if conn:
            conn.rollback()
        return HTMLResponse(f"<h1>DB ERROR</h1><pre>{e}</pre>", status_code=500)

    finally:
        safe_close(conn, cur)

    return RedirectResponse("/all_new", status_code=303)

# ================= PRODUCTS =================
@app.get("/all", response_class=HTMLResponse)
def all_products(
    request: Request,
    auth: str = Cookie(default=None),
    mode: str = Cookie(default="driver"),
    q: str = None
):

    if auth != "ok":
        return RedirectResponse("/login", status_code=303)

    conn = get_conn()
    cur = conn.cursor()

    if q:
        cur.execute("""
            SELECT code,name,manufacturer,quantity,min_limit
            FROM products
            WHERE code ILIKE %s
            OR manufacturer ILIKE %s
            ORDER BY manufacturer,name
        """, (f"%{q}%", f"%{q}%"))
    else:
        cur.execute("""
            SELECT code,name,manufacturer,quantity,min_limit
            FROM products
            ORDER BY manufacturer,name
        """)

    rows = cur.fetchall()
    safe_close(conn, cur)

    from collections import defaultdict
    grouped = defaultdict(list)

    for row in rows:
        grouped[row[2] or "Neznámý"].append({
            "code": row[0],
            "name": row[1],
            "manufacturer": row[2] or "Neznámý",
            "quantity": row[3],
            "min_limit": row[4]
        })

    user = request.cookies.get("user", "Neznámý")
    mode_label = "SKLAD" if mode == "sklad" else "ŘIDIČ"

    return templates.TemplateResponse(
        "all_new.html",
        {
            "request": request,
            "title": "Sklad",
            "user": user,
            "mode": mode_label,
            "grouped": grouped,
            "mode_raw": mode
        }
    )

# ================= LOW =================
@app.get("/low", response_class=HTMLResponse)
def low(request: Request, auth: str = Cookie(default=None)):

    if auth != "ok":
        return RedirectResponse("/login", status_code=303)

    conn = get_conn()
    cur = conn.cursor()

    cur.execute("""
    SELECT code,name,manufacturer,quantity,min_limit
    FROM products
    WHERE quantity <= min_limit
    ORDER BY manufacturer,name
    """)

    rows = cur.fetchall()
    safe_close(conn, cur)

    from collections import defaultdict
    grouped = defaultdict(list)

    for row in rows:

        grouped[row[2] or "Neznámý"].append({
            "code": row[0],
            "name": row[1],
            "manufacturer": row[2] or "Neznámý",
            "quantity": row[3],
            "min_limit": row[4]
        })

    user = request.cookies.get("user", "Neznámý")
    mode = request.cookies.get("mode", "driver")

    return templates.TemplateResponse(
        "low_new.html",
        {
            "request": request,
            "title": "Nízký stav",
            "user": user,
            "mode": mode,
            "grouped": grouped
        }
    )

# ================= HISTORY =================
@app.get("/history", response_class=HTMLResponse)
def history(request: Request, auth: str = Cookie(default=None)):

    if auth != "ok":
        return RedirectResponse("/login", status_code=303)

    conn = get_conn()
    cur = conn.cursor()

    cur.execute("""
        SELECT m.code,
               p.name,
               m.change,
               m.created_at,
               COALESCE(m.user_name,'Unknown')
        FROM movements m
        LEFT JOIN products p ON p.code = m.code
        ORDER BY m.created_at DESC
        LIMIT 200
    """)

    data = cur.fetchall()

    safe_close(conn, cur)

    rows=[]

    for r in data:

        rows.append({
            "code": r[0],
            "name": r[1] or "-",
            "change": r[2],
            "date": r[3],
            "user": r[4]
        })

    user = request.cookies.get("user", "Neznámý")
    mode = request.cookies.get("mode", "driver")

    return templates.TemplateResponse(
        "history_new.html",
        {
            "request": request,
            "title": "Historie",
            "user": user,
            "mode": mode,
            "rows": rows
        }
    )

@app.get("/all_new", response_class=HTMLResponse)
def all_new(request: Request, auth: str = Cookie(default=None)):

    if auth != "ok":
        return RedirectResponse("/login", status_code=303)

    conn = get_conn()
    cur = conn.cursor()

    cur.execute("""
        SELECT code,name,manufacturer,quantity,min_limit
        FROM products
        ORDER BY manufacturer,name
    """)

    rows = cur.fetchall()
    safe_close(conn, cur)

    from collections import defaultdict
    grouped = defaultdict(list)

    for row in rows:
        grouped[row[2] or "Neznámý"].append({
            "code": row[0],
            "name": row[1],
            "manufacturer": row[2] or "Neznámý",
            "quantity": row[3],
            "min_limit": row[4]
        })

    user = request.cookies.get("user", "Neznámý")
    mode = request.cookies.get("mode", "driver")
    users = get_app_users()

    return templates.TemplateResponse(
        "all_new.html",
        {
            "request": request,
            "title": "Sklad",
            "user": user,
            "mode": mode,
            "grouped": grouped,
            "users": users
        }
    )

from fastapi import Request

@app.post("/set_quantity")
async def set_quantity(request: Request):

    data = await request.json()

    code = data["code"]
    qty = int(data["qty"])

    conn = get_conn()
    cur = conn.cursor()

    cur.execute(
        "UPDATE products SET quantity=%s WHERE code=%s",
        (qty, code)
    )

    conn.commit()

    safe_close(conn, cur)

    return {"status": "ok"}

@app.post("/upload_login_background")
async def upload_login_background(
    auth: str = Cookie(default=None),
    photo: UploadFile = File(...)
):
    if auth != "ok":
        return RedirectResponse("/login", status_code=303)

    folder = "static"
    os.makedirs(folder, exist_ok=True)

    filepath = os.path.join(folder, "login_bg.jpg")

    with open(filepath, "wb") as buffer:
        buffer.write(await photo.read())

    conn = None
    cur = None
    try:
        conn = get_conn()
        cur = conn.cursor()
        cur.execute("""
            UPDATE app_settings
            SET login_background='/static/login_bg.jpg',
                updated_at=CURRENT_TIMESTAMP
            WHERE id = (SELECT id FROM app_settings ORDER BY id LIMIT 1)
        """)
        conn.commit()
    finally:
        safe_close(conn, cur)

    return RedirectResponse("/select_user?edit=1", status_code=303)

@app.post("/upload_photo")
async def upload_photo(code: str = Form(...), photo: UploadFile = File(...)):

    folder = "product_images"
    os.makedirs(folder, exist_ok=True)

    filepath = f"{folder}/{code}.jpg"

    with open(filepath, "wb") as buffer:
        buffer.write(await photo.read())

    return {"status": "ok"}

@app.get("/product_img/{code}")
def get_product_img(code: str):

    path = f"product_images/{code}.jpg"

    if os.path.exists(path):
        return FileResponse(path)

    return {"error": "no image"}